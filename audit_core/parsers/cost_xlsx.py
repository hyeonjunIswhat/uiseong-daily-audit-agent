"""SW개발비 산출내역서(xlsx) 검산기 — KOSA 대가산정 가이드 표준 양식 대응.

실물 근거(2026 생성형AI플랫폼 산출내역서, 검토요청·조치반영 2개 버전):
  'SW개발비 산정' 시트에 ①개발원가(총FP × 단가 × 보정계수들) ②이윤(원가×율)
  ③직접경비(내역표 합산) ④부가세 ⑤최종액('부가세 포함, 십만단위 절사' 표기)
  구조. FP 곱셈은 셀에 분리되어 × 기호가 없으므로 텍스트 검산이 불가 — 셀 좌표
  구조를 직접 읽어 재계산한다.

보수 원칙(오탐 방지):
  - 서식을 인식하지 못하면 CostSheetError — 임의 재구성하지 않는다
  - 직접경비 내역식은 단순 곱셈 표기('150,000원*3인*10회')만 검산, 복합 표기는 건너뜀
  - '절사' 표기가 있는 최종액은 1,000,000원 미만의 하향 차이를 허용(예산 맞춤 절사 관행)
  - 이윤율 25% 초과는 별도 FLAG (SW사업 대가 기준 상한)
"""

import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from audit_core.agents.verifier import (
    _FRACTION_TOL_ABS,
    _FRACTION_TOL_REL,
    NumericCheck,
)

_NUM_TOKEN = re.compile(r"[0-9][0-9,]*")
PROFIT_RATE_CAP = Decimal("0.25")
TRUNC_ALLOWANCE = 1_000_000  # '절사' 표기 시 허용 하향 차이


class CostSheetError(Exception):
    pass


@dataclass
class CostSheetResult:
    sheet: str
    checks: list[NumericCheck] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def flags(self) -> list[NumericCheck]:
        return [c for c in self.checks if not c.match]


def _cells(ws, row: int) -> list:
    return [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]


def _numerics(values) -> list[Decimal]:
    return [Decimal(str(v)) for v in values if isinstance(v, (int, float))]


def _find_row(ws, keyword: str, col_a_only: bool = True) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and keyword in v:
            return r
        if not col_a_only:
            for c in range(2, ws.max_column + 1):
                v2 = ws.cell(r, c).value
                if isinstance(v2, str) and keyword in v2:
                    return r
    return None


def _approx(expected: Decimal, claimed: Decimal) -> bool:
    tol = max(_FRACTION_TOL_ABS, int(expected * _FRACTION_TOL_REL))
    return abs(int(expected) - int(claimed)) <= tol


def check_cost_sheet(path: str | Path) -> CostSheetResult:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheet = next((n for n in wb.sheetnames if "개발비" in n or "산정" in n), wb.sheetnames[0])
    ws = wb[sheet]
    res = CostSheetResult(sheet=sheet)

    # ① 개발원가 = 총FP × 단가 × 보정계수들 (헤더 '총기능점수' 아래 첫 숫자 행)
    hdr = _find_row(ws, "총기능점수")
    if hdr is None:
        raise CostSheetError("'총기능점수' 헤더를 찾지 못함 — 표준 양식 아님")
    dev_cost = None
    for r in range(hdr + 1, min(hdr + 5, ws.max_row) + 1):
        nums = _numerics(_cells(ws, r))
        if len(nums) >= 4:
            *factors, claimed = nums
            expected = Decimal(1)
            for f in factors:
                expected *= f
            res.checks.append(NumericCheck(
                "fp_dev_cost", f"개발원가 = {' × '.join(str(f) for f in factors)}",
                int(expected), int(claimed), _approx(expected, claimed)))
            dev_cost = claimed
            break
    if dev_cost is None:
        raise CostSheetError("개발원가 산정 행(FP·단가·계수)을 찾지 못함")

    # ② 이윤 = 개발원가 × 이윤율 (상한 25%)
    profit = Decimal(0)
    r = _find_row(ws, "이윤")
    if r:
        nums = _numerics(_cells(ws, r))
        if len(nums) >= 2:
            rate, amount = nums[-2], nums[-1]
            expected = dev_cost * rate
            res.checks.append(NumericCheck(
                "fp_profit", f"이윤 = 개발원가 × {rate}",
                int(expected), int(amount), _approx(expected, amount)))
            if rate > PROFIT_RATE_CAP:
                res.checks.append(NumericCheck(
                    "fp_profit_rate", f"이윤율 {rate} — 상한 25% 초과",
                    int(PROFIT_RATE_CAP * 100), int(rate * 100), False))
            profit = amount

    # ③ 직접경비 — 요약행 vs 내역표 '합 계' 교차 + 내역 단순곱셈 검산
    direct = Decimal(0)
    r_sum = _find_row(ws, "직접경비")
    if r_sum:
        nums = _numerics(_cells(ws, r_sum))
        if nums:
            direct = nums[-1]
    r_detail = _find_row(ws, "○ 직접경비", col_a_only=False)
    if r_detail:
        detail_total = None
        for r in range(r_detail + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if isinstance(label, str) and label.replace(" ", "").startswith("합계"):
                nums = _numerics(_cells(ws, r))
                detail_total = nums[-1] if nums else None
                break
            expr = next((ws.cell(r, c).value for c in range(2, ws.max_column + 1)
                         if isinstance(ws.cell(r, c).value, str) and "*" in ws.cell(r, c).value), None)
            nums = _numerics(_cells(ws, r))
            # 단순 곱셈 내역만 검산 — 복합 표기(줄바꿈·콜론·단위환산)는 보수적으로 건너뜀
            if expr and nums and "\n" not in expr and ":" not in expr:
                tokens = _NUM_TOKEN.findall(expr)
                if len(tokens) >= 2:
                    prod = Decimal(1)
                    for t in tokens:
                        prod *= Decimal(t.replace(",", ""))
                    res.checks.append(NumericCheck(
                        "fp_direct_item", f"{str(ws.cell(r, 1).value or '').strip()}: {expr.strip()}",
                        int(prod), int(nums[-1]), int(prod) == int(nums[-1])))
        if detail_total is not None and direct:
            res.checks.append(NumericCheck(
                "fp_direct_total", "직접경비 요약 = 내역표 합계",
                int(detail_total), int(direct), int(detail_total) == int(direct)))

    # ④ 부가세 = (개발원가+이윤+직접경비) × 세율
    vat = Decimal(0)
    r = _find_row(ws, "부가세")
    if r:
        nums = _numerics(_cells(ws, r))
        if len(nums) >= 2:
            rate, amount = nums[-2], nums[-1]
            expected = (dev_cost + profit + direct) * rate
            res.checks.append(NumericCheck(
                "fp_vat", f"부가세 = (개발원가+이윤+직접경비) × {rate}",
                int(expected), int(amount), _approx(expected, amount)))
            vat = amount

    # ⑤ 최종 SW개발비 — '절사' 표기 시 1,000,000원 미만 하향 차이 허용
    r = _find_row(ws, "개발비(부가세") or _find_row(ws, "소프트웨어 개발비")
    if r:
        label = str(ws.cell(r, 1).value or "")
        nums = _numerics(_cells(ws, r))
        if nums:
            claimed = nums[-1]
            expected = dev_cost + profit + direct + (vat if "부가세" in label else Decimal(0))
            if "절사" in label:
                ok = Decimal(0) <= expected - claimed < TRUNC_ALLOWANCE
                if ok and expected - claimed > 0:
                    res.notes.append(
                        f"최종액 절사 차이 {int(expected - claimed):,}원 (표기된 절사 관행 — 허용)")
            else:
                ok = _approx(expected, claimed)
            res.checks.append(NumericCheck(
                "fp_total", label.strip(), int(expected), int(claimed), ok))

    return res
